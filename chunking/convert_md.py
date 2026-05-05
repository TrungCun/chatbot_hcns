import os
os.environ["CUDA_VISIBLE_DEVICES"] = "1"  # must be set before torch import

import re
import json
import logging
import traceback
import subprocess
from datetime import datetime
from pathlib import Path

import openpyxl
import pdfplumber
import pytesseract
import torch
from docling.document_converter import DocumentConverter as DoclingConverter
from pdf2image import convert_from_path
from PIL import Image
from transformers import AutoModel, AutoTokenizer
from docx import Document as DocxDocument

logging.basicConfig(level=logging.INFO, format="%(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

SUPPORTED_EXT = {".pdf", ".docx", ".doc", ".xlsx", ".xls", ".pptx", ".ppt", ".md", ".txt"}

PDF_OCR_STYLE = {
    "prompt": "<image>\n<|grounding|>Convert the document to markdown.",
    "base_size": 1024,
    "image_size": 640,
    "crop_mode": True,
    "test_compress": True,
    "save_results": True,
}


class FileConverter:
    """
    Tool chuyển đổi tài liệu (PDF, DOCX, XLSX, ...) sang Markdown.


    Ví dụ:
        converter = FileConverter(model_id="...")
        result  = converter.convert_file("doc.pdf", output_dir="/tmp")
        summary = converter.process_folder("/data/input")
        converter.cleanup()

    Hoặc dùng context manager:
        with FileConverter(model_id="...") as converter:
            converter.process_folder(...)
    """

    def __init__(self, model_id: str):
        logger.info("Loading tokenizer from %s", model_id)
        self.tokenizer = AutoTokenizer.from_pretrained(model_id, trust_remote_code=True)
        self._patch_llama_compat()
        logger.info("Loading model from %s", model_id)
        self.model = AutoModel.from_pretrained(
            model_id,
            trust_remote_code=True,
            use_safetensors=True,
            device_map="auto",
            torch_dtype=torch.bfloat16,
        ).eval()
        self.docling = DoclingConverter()

    def __enter__(self):
        return self

    def __exit__(self, *_):
        self.cleanup()


    def convert_file(self, file_path: str, output_dir: str) -> dict:
        """
        Chuyển đổi một file sang Markdown.
        Returns: {"file": str, "success": bool, "error": str | None}
        """
        file_path = str(file_path)
        ext = Path(file_path).suffix.lower()
        try:
            if ext == ".doc":
                file_path = str(self._doc_to_docx(file_path))
                ext = ".docx"

            if ext == ".pdf":
                content = (
                    self._pdf_scan(file_path, output_dir)
                    if self._is_pdf_scan(file_path)
                    else self._docling(file_path)
                )
            else:
                content = self._non_pdf(file_path, ext)

            if not content:
                raise ValueError("Empty content after conversion")

            Path(file_path).with_suffix(".md").write_text(content, encoding="utf-8")
            logger.info("OK: %s", Path(file_path).name)
            return {"file": file_path, "success": True, "error": None}

        except Exception as exc:
            traceback.print_exc()
            logger.warning("Failed: %s - %s", file_path, exc)
            return {"file": file_path, "success": False, "error": str(exc)}

    def cleanup(self):
        """Giải phóng bộ nhớ GPU."""
        del self.model
        torch.cuda.empty_cache()

    # ─────────────────────────── INTERNAL ───────────────────────────

    def _patch_llama_compat(self):
        try:
            from transformers.models.llama import modeling_llama
            if not hasattr(modeling_llama, "LlamaFlashAttention2"):
                setattr(modeling_llama, "LlamaFlashAttention2", modeling_llama.LlamaAttention)
        except ImportError:
            pass

    def _auto_rotate(self, image_path: Path) -> None:
        try:
            img = Image.open(image_path)
            osd = pytesseract.image_to_osd(img, output_type=pytesseract.Output.DICT)
            angle = osd.get("rotate", 0)
            if angle:
                img.rotate(-angle, expand=True).save(image_path)
        except Exception:
            pass

    def _is_pdf_scan(self, pdf_path: str) -> bool:
        try:
            total_chars = 0
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        total_chars += len(text.strip())
                        if total_chars >= 100:
                            return False
            return True
        except Exception as exc:
            logger.warning("Unable to check PDF scan status: %s", exc)
            return True

    def _clean_empty_tables(self, text: str) -> str:
        keywords = {"lần", "số yêu cầu", "sửa đổi", "nội dung", "ngày", "người",
                    "revision", "date", "description", "author"}
        tables = re.findall(r"<table.*?>.*?</table>", text, flags=re.DOTALL | re.IGNORECASE)
        for table in tables:
            cells = re.findall(r"<td.*?>\s*(.*?)\s*</td>", table, flags=re.DOTALL | re.IGNORECASE)
            cleaned = [re.sub(r"<.*?>", "", c).strip() for c in cells]
            has_data = any(c and not any(kw in c.lower() for kw in keywords) for c in cleaned)
            if not has_data:
                text = text.replace(table, "")
        return re.sub(r"\n{3,}", "\n\n", text).strip()

    def _pdf_scan(self, pdf_path: str, output_dir: str) -> str:
        pages = convert_from_path(pdf_path, dpi=300)
        os.makedirs(output_dir, exist_ok=True)
        output = []
        for i, page in enumerate(pages):
            img_path = Path(output_dir) / f"temp_{i}.png"
            try:
                page.save(img_path)
                self._auto_rotate(img_path)
                res = self.model.infer(
                    self.tokenizer,
                    prompt=PDF_OCR_STYLE["prompt"],
                    image_file=str(img_path),
                    output_path=str(output_dir),
                    base_size=PDF_OCR_STYLE["base_size"],
                    image_size=PDF_OCR_STYLE["image_size"],
                    crop_mode=PDF_OCR_STYLE["crop_mode"],
                    save_results=PDF_OCR_STYLE["save_results"],
                    test_compress=PDF_OCR_STYLE["test_compress"],
                )
                text = res["text"] if isinstance(res, dict) else str(res)
                output.append(f"\n## Page {i + 1}\n{self._clean_empty_tables(text)}")
            finally:
                img_path.unlink(missing_ok=True)
        return "\n".join(output)

    def _docling(self, file_path: str) -> str:
        return self.docling.convert(file_path).document.export_to_markdown()

    def _non_pdf(self, file_path: str, ext: str) -> str:
        try:
            return self._docling(file_path)
        except Exception as exc:
            logger.warning("Docling failed for %s, falling back: %s", file_path, exc)
            if ext == ".docx":
                return self._docx(file_path)
            if ext in {".xlsx", ".xls"}:
                return self._xlsx(file_path)
            raise

    def _docx(self, file_path: str) -> str:
        doc = DocxDocument(file_path)
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())

    def _doc_to_docx(self, doc_path: str) -> Path:
        dest = Path(doc_path).with_suffix(".docx")
        if dest.exists():
            return dest
        try:
            subprocess.run(
                ["soffice", "--headless", "--convert-to", "docx", "--outdir", str(dest.parent), str(doc_path)],
                check=True,
                capture_output=True,
            )
        except subprocess.CalledProcessError as err:
            msg = err.stderr.decode(errors="ignore") if err.stderr else "conversion failed"
            raise RuntimeError(f"Unable to convert {doc_path} to docx: {msg}")
        if not dest.exists():
            raise RuntimeError(f"Converted file missing: {dest}")
        return dest

    def _xlsx(self, file_path: str) -> str:
        wb = openpyxl.load_workbook(file_path)
        content = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            content.append(f"\n## Sheet: {sheet}")
            for row in ws.iter_rows(values_only=True):
                content.append(" | ".join(str(c) if c else "" for c in row))
        return "\n".join(content)


# ─────────────────────────── ENTRY POINT ───────────────────────────

def main():
    model_id = "/home/trung-ai/chatbot_hcns/weights/DeepSeek-OCR-bnb-4bit-NF4"
    file_path = "/home/trung-ai/chatbot_hcns/124ebf20b3df468995e5dd4829d1b357 (1).pdf"

    with FileConverter(model_id) as converter:
        converter.convert_file(file_path, output_dir="/tmp")

if __name__ == "__main__":
    main()